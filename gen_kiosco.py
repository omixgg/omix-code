#!/usr/bin/env python3
"""Genera planilla Excel para gestion de kiosco: costos, facturacion, beneficios."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

wb = openpyxl.Workbook()

# ── Estilos reutilizables ─────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CURRENCY_FMT = '#,##0.00" $"'
PCT_FMT = '0.0%'
DATE_FMT = 'DD/MM/YYYY'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def style_header(ws, cols, row=1):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

def auto_width(ws, cols, min_w=10, max_w=40):
    for c in range(1, cols + 1):
        letter = get_column_letter(c)
        best = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 3, max_w))
        ws.column_dimensions[letter].width = best

# ═══════════════════════════════════════════════════════════════════════
# HOJA 1: PRODUCTOS (catalogo)
# ═══════════════════════════════════════════════════════════════════════
ws_prod = wb.active
ws_prod.title = "Productos"
ws_prod.sheet_properties.tabColor = "2F5496"

prod_headers = ["Codigo", "Producto", "Categoria", "Proveedor", "Costo Unit.",
                "Precio Venta", "Margen %", "Stock Min.", "Stock Actual", "Estado"]
for i, h in enumerate(prod_headers, 1):
    ws_prod.cell(row=1, column=i, value=h)
style_header(ws_prod, len(prod_headers))

# Datos de ejemplo
productos_ej = [
    ["P001", "Coca-Cola 500ml", "Bebidas", "Distribuidora XYZ", 850.00, 1200.00, None, 12, 24, "OK"],
    ["P002", "Fanta 500ml", "Bebidas", "Distribuidora XYZ", 780.00, 1100.00, None, 10, 8, "REPONER"],
    ["P003", "Alfajor Jorgito", "Golosinas", "Mayorista ABC", 350.00, 600.00, None, 20, 45, "OK"],
    ["P004", "Papas Lays 150g", "Snacks", "Mayorista ABC", 1100.00, 1800.00, None, 8, 15, "OK"],
    ["P005", "Marlboro Box", "Cigarrillos", "Tabacalera SRL", 1800.00, 2200.00, None, 5, 3, "REPONER"],
    ["P006", "Agua Mineral 1L", "Bebidas", "Distribuidora XYZ", 400.00, 700.00, None, 15, 30, "OK"],
    ["P007", "Chocolate Milka", "Golosinas", "Mayorista ABC", 1200.00, 2000.00, None, 6, 10, "OK"],
    ["P008", "Cerveza Quilmes 1L", "Bebidas", "Distribuidora XYZ", 950.00, 1500.00, None, 12, 6, "REPONER"],
    ["P009", "Galletitas Oreo", "Golosinas", "Mayorista ABC", 650.00, 1000.00, None, 10, 20, "OK"],
    ["P010", "Hielo 2kg", "Otros", "Fabrica Hielo SA", 500.00, 900.00, None, 15, 12, "OK"],
]

for r, prod in enumerate(productos_ej, 2):
    for c, val in enumerate(prod, 1):
        ws_prod.cell(row=r, column=c, value=val)

# Formula Margen %: (Precio Venta - Costo) / Precio Venta
last_prod = len(productos_ej) + 1
for r in range(2, last_prod + 1):
    ws_prod.cell(row=r, column=7).value = f'=(F{r}-E{r})/F{r}'
    ws_prod.cell(row=r, column=7).number_format = PCT_FMT

# Formato moneda
for r in range(2, last_prod + 1):
    ws_prod.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws_prod.cell(row=r, column=6).number_format = CURRENCY_FMT

# Formato condicional: REPONER en rojo
ws_prod.conditional_formatting.add(
    f"J2:J{last_prod + 50}",
    CellIsRule(operator="equal", formula=['"REPONER"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
)
ws_prod.conditional_formatting.add(
    f"J2:J{last_prod + 50}",
    CellIsRule(operator="equal", formula=['"OK"'], fill=GREEN_FILL)
)

# Barra de datos en Stock Actual
ws_prod.conditional_formatting.add(
    f"I2:I{last_prod + 50}",
    DataBarRule(start_type="min", end_type="max", color="2F5496", showValue=True)
)

style_data_rows(ws_prod, 2, last_prod, len(prod_headers))
auto_width(ws_prod, len(prod_headers))

# ═══════════════════════════════════════════════════════════════════════
# HOJA 2: COMPRAS (costos)
# ═══════════════════════════════════════════════════════════════════════
ws_comp = wb.create_sheet("Compras")
ws_comp.sheet_properties.tabColor = "ED7D31"

comp_headers = ["Fecha", "Codigo", "Producto", "Proveedor", "Cantidad",
                "Costo Unit.", "Costo Total", "Forma Pago", "Nro. Factura", "Notas"]
for i, h in enumerate(comp_headers, 1):
    ws_comp.cell(row=1, column=i, value=h)
style_header(ws_comp, len(comp_headers))

compras_ej = [
    [date(2025, 5, 1), "P001", "Coca-Cola 500ml", "Distribuidora XYZ", 24, 850.00, None, "Efectivo", "F-001", ""],
    [date(2025, 5, 1), "P002", "Fanta 500ml", "Distribuidora XYZ", 12, 780.00, None, "Efectivo", "F-001", ""],
    [date(2025, 5, 3), "P003", "Alfajor Jorgito", "Mayorista ABC", 50, 340.00, None, "Transferencia", "F-045", ""],
    [date(2025, 5, 5), "P005", "Marlboro Box", "Tabacalera SRL", 10, 1750.00, None, "Efectivo", "F-102", "Aumento proveedor"],
    [date(2025, 5, 7), "P008", "Cerveza Quilmes 1L", "Distribuidora XYZ", 18, 920.00, None, "Transferencia", "F-003", "Promo 10% desc"],
    [date(2025, 5, 10), "P004", "Papas Lays 150g", "Mayorista ABC", 20, 1080.00, None, "Efectivo", "F-050", ""],
    [date(2025, 5, 12), "P006", "Agua Mineral 1L", "Distribuidora XYZ", 36, 380.00, None, "Transferencia", "F-004", "Compra x mayor"],
]

for r, comp in enumerate(compras_ej, 2):
    for c, val in enumerate(comp, 1):
        ws_comp.cell(row=r, column=c, value=val)

last_comp = len(compras_ej) + 1
for r in range(2, last_comp + 1):
    ws_comp.cell(row=r, column=7).value = f'=E{r}*F{r}'  # Costo Total = Cant * Unitario
    ws_comp.cell(row=r, column=1).number_format = DATE_FMT
    ws_comp.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws_comp.cell(row=r, column=7).number_format = CURRENCY_FMT

style_data_rows(ws_comp, 2, last_comp, len(comp_headers))
auto_width(ws_comp, len(comp_headers))

# ═══════════════════════════════════════════════════════════════════════
# HOJA 3: VENTAS (facturacion)
# ═══════════════════════════════════════════════════════════════════════
ws_vent = wb.create_sheet("Ventas")
ws_vent.sheet_properties.tabColor = "4472C4"

vent_headers = ["Fecha", "Codigo", "Producto", "Cantidad", "Precio Unit.",
                "Total Venta", "Costo Unit.", "Costo Total", "Beneficio",
                "Margen %", "Forma Pago", "Notas"]
for i, h in enumerate(vent_headers, 1):
    ws_vent.cell(row=1, column=i, value=h)
style_header(ws_vent, len(vent_headers))

ventas_ej = [
    [date(2025, 5, 2), "P001", "Coca-Cola 500ml", 15, 1200.00, None, 850.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 2), "P003", "Alfajor Jorgito", 8, 600.00, None, 350.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 3), "P005", "Marlboro Box", 4, 2200.00, None, 1800.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 4), "P001", "Coca-Cola 500ml", 10, 1200.00, None, 850.00, None, None, None, "MercadoPago", ""],
    [date(2025, 5, 5), "P004", "Papas Lays 150g", 6, 1800.00, None, 1100.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 6), "P008", "Cerveza Quilmes 1L", 12, 1500.00, None, 950.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 7), "P006", "Agua Mineral 1L", 20, 700.00, None, 400.00, None, None, None, "MercadoPago", ""],
    [date(2025, 5, 8), "P007", "Chocolate Milka", 3, 2000.00, None, 1200.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 9), "P002", "Fanta 500ml", 5, 1100.00, None, 780.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 10), "P001", "Coca-Cola 500ml", 12, 1200.00, None, 850.00, None, None, None, "MercadoPago", ""],
    [date(2025, 5, 10), "P009", "Galletitas Oreo", 7, 1000.00, None, 650.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 11), "P010", "Hielo 2kg", 8, 900.00, None, 500.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 12), "P003", "Alfajor Jorgito", 15, 600.00, None, 350.00, None, None, None, "Efectivo", ""],
    [date(2025, 5, 13), "P008", "Cerveza Quilmes 1L", 6, 1500.00, None, 950.00, None, None, None, "MercadoPago", ""],
    [date(2025, 5, 14), "P005", "Marlboro Box", 3, 2200.00, None, 1800.00, None, None, None, "Efectivo", ""],
]

for r, venta in enumerate(ventas_ej, 2):
    for c, val in enumerate(venta, 1):
        ws_vent.cell(row=r, column=c, value=val)

last_vent = len(ventas_ej) + 1
for r in range(2, last_vent + 1):
    ws_vent.cell(row=r, column=6).value = f'=D{r}*E{r}'      # Total Venta
    ws_vent.cell(row=r, column=8).value = f'=D{r}*G{r}'      # Costo Total
    ws_vent.cell(row=r, column=9).value = f'=F{r}-H{r}'      # Beneficio
    ws_vent.cell(row=r, column=10).value = f'=I{r}/F{r}'     # Margen %
    # Formatos
    ws_vent.cell(row=r, column=1).number_format = DATE_FMT
    for col in [5, 6, 7, 8, 9]:
        ws_vent.cell(row=r, column=col).number_format = CURRENCY_FMT
    ws_vent.cell(row=r, column=10).number_format = PCT_FMT

# Formato condicional: Beneficio positivo verde, negativo rojo
ws_vent.conditional_formatting.add(
    f"I2:I{last_vent + 100}",
    CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL)
)
ws_vent.conditional_formatting.add(
    f"I2:I{last_vent + 100}",
    CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
)

style_data_rows(ws_vent, 2, last_vent, len(vent_headers))
auto_width(ws_vent, len(vent_headers))

# ═══════════════════════════════════════════════════════════════════════
# HOJA 4: DASHBOARD (resumen)
# ═══════════════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = "70AD47"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="2F5496")
SUB_FONT = Font(name="Calibri", size=12, bold=True, color="2F5496")
LABEL_FONT = Font(name="Calibri", size=11, bold=True)

# ── Seccion 1: KPIs ──
ws_dash.merge_cells("A1:F1")
ws_dash.cell(row=1, column=1, value="DASHBOARD - KIOSCO").font = TITLE_FONT
ws_dash.cell(row=1, column=1).alignment = Alignment(horizontal="center")

ws_dash.merge_cells("A2:F2")
ws_dash.cell(row=2, column=1, value=f"Actualizado: {date.today().strftime('%d/%m/%Y')}").font = Font(name="Calibri", size=10, italic=True, color="666666")
ws_dash.cell(row=2, column=1).alignment = Alignment(horizontal="center")

kpis = [
    ("A4", "Ventas Totales", "=SUM(Ventas!F:F)"),
    ("B4", "Costos Totales", "=SUM(Ventas!H:H)"),
    ("C4", "Beneficio Bruto", "=SUM(Ventas!I:I)"),
    ("D4", "Margen Promedio", "=IF(B4>0,C4/B4,0)"),
    ("E4", "Transacciones", "=COUNTA(Ventas!A2:A2000)"),
    ("F4", "Ticket Promedio", "=IF(E4>0,A4/E4,0)"),
]

for cell, label, formula in kpis:
    r = 4
    ws_dash.cell(row=r, column=ord(cell[0]) - 64, value=label).font = LABEL_FONT
    ws_dash.cell(row=r, column=ord(cell[0]) - 64).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws_dash.cell(row=r, column=ord(cell[0]) - 64).border = THIN_BORDER
    ws_dash.cell(row=r, column=ord(cell[0]) - 64).alignment = Alignment(horizontal="center")

    ws_dash.cell(row=r + 1, column=ord(cell[0]) - 64, value=formula)
    ws_dash.cell(row=r + 1, column=ord(cell[0]) - 64).font = Font(name="Calibri", size=18, bold=True, color="2F5496")
    ws_dash.cell(row=r + 1, column=ord(cell[0]) - 64).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=r + 1, column=ord(cell[0]) - 64).border = THIN_BORDER
    # Formato moneda en KPIs de dinero
    if cell in ("A4", "B4", "C4", "F4"):
        ws_dash.cell(row=r + 1, column=ord(cell[0]) - 64).number_format = CURRENCY_FMT
    elif cell == "D4":
        ws_dash.cell(row=r + 1, column=ord(cell[0]) - 64).number_format = PCT_FMT

# ── Seccion 2: Ventas por producto ──
ws_dash.cell(row=7, column=1, value="TOP PRODUCTOS POR VENTA").font = SUB_FONT
ws_dash.merge_cells("A7:C7")

ws_dash.cell(row=8, column=1, value="Producto").font = HEADER_FONT
ws_dash.cell(row=8, column=1).fill = HEADER_FILL
ws_dash.cell(row=8, column=2, value="Cantidad Vendida").font = HEADER_FONT
ws_dash.cell(row=8, column=2).fill = HEADER_FILL
ws_dash.cell(row=8, column=3, value="Total Ventas").font = HEADER_FONT
ws_dash.cell(row=8, column=3).fill = HEADER_FILL

# SUMAR.SI por producto (10 filas)
productos_unicos = ["Coca-Cola 500ml", "Fanta 500ml", "Alfajor Jorgito", "Papas Lays 150g",
                    "Marlboro Box", "Agua Mineral 1L", "Chocolate Milka", "Cerveza Quilmes 1L",
                    "Galletitas Oreo", "Hielo 2kg"]
for i, prod in enumerate(productos_unicos):
    r = 9 + i
    ws_dash.cell(row=r, column=1, value=prod).border = THIN_BORDER
    ws_dash.cell(row=r, column=2, value=f'=SUMIF(Ventas!C:C,A{r},Ventas!D:D)').border = THIN_BORDER
    ws_dash.cell(row=r, column=3, value=f'=SUMIF(Ventas!C:C,A{r},Ventas!F:F)').border = THIN_BORDER
    ws_dash.cell(row=r, column=3).number_format = CURRENCY_FMT

# ── Seccion 3: Ventas por dia ──
ws_dash.cell(row=7, column=5, value="VENTAS POR DIA").font = SUB_FONT

ws_dash.cell(row=8, column=5, value="Fecha").font = HEADER_FONT
ws_dash.cell(row=8, column=5).fill = HEADER_FILL
ws_dash.cell(row=8, column=6, value="Total").font = HEADER_FONT
ws_dash.cell(row=8, column=6).fill = HEADER_FILL

dias_ej = [
    date(2025, 5, 2), date(2025, 5, 3), date(2025, 5, 4), date(2025, 5, 5),
    date(2025, 5, 6), date(2025, 5, 7), date(2025, 5, 8), date(2025, 5, 9),
    date(2025, 5, 10), date(2025, 5, 11), date(2025, 5, 12), date(2025, 5, 13), date(2025, 5, 14),
]
for i, d in enumerate(dias_ej):
    r = 9 + i
    ws_dash.cell(row=r, column=5, value=d).border = THIN_BORDER
    ws_dash.cell(row=r, column=5).number_format = DATE_FMT
    ws_dash.cell(row=r, column=6, value=f'=SUMIF(Ventas!A:A,E{r},Ventas!F:F)').border = THIN_BORDER
    ws_dash.cell(row=r, column=6).number_format = CURRENCY_FMT

# ── Grafico de barras: Ventas por producto ──
chart = BarChart()
chart.type = "col"
chart.title = "Ventas por Producto ($)"
chart.y_axis.title = "Total Ventas"
chart.x_axis.title = "Producto"
chart.style = 10
chart.width = 22
chart.height = 12

data_ref = Reference(ws_dash, min_col=3, min_row=8, max_row=18)
cats_ref = Reference(ws_dash, min_col=1, min_row=9, max_row=18)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "2F5496"
ws_dash.add_chart(chart, "B22")

# ── Grafico torta: Metodo de pago ──
ws_dash.cell(row=20, column=5, value="METODO DE PAGO").font = SUB_FONT

metodos = ["Efectivo", "MercadoPago"]
ws_dash.cell(row=21, column=5, value="Metodo").font = HEADER_FONT
ws_dash.cell(row=21, column=5).fill = HEADER_FILL
ws_dash.cell(row=21, column=6, value="Total").font = HEADER_FONT
ws_dash.cell(row=21, column=6).fill = HEADER_FILL
for i, m in enumerate(metodos):
    r = 22 + i
    ws_dash.cell(row=r, column=5, value=m).border = THIN_BORDER
    ws_dash.cell(row=r, column=6, value=f'=SUMIF(Ventas!K:K,E{r},Ventas!F:F)').border = THIN_BORDER
    ws_dash.cell(row=r, column=6).number_format = CURRENCY_FMT

pie = PieChart()
pie.title = "Ventas por Medio de Pago"
pie.width = 16
pie.height = 12
pie_data = Reference(ws_dash, min_col=6, min_row=21, max_row=23)
pie_cats = Reference(ws_dash, min_col=5, min_row=22, max_row=23)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
ws_dash.add_chart(pie, "E24")

auto_width(ws_dash, 6)

# ── Proteger formulas ──
ws_dash.sheet_protection.sheet = True
ws_dash.sheet_protection.set_password("kiosco")

# ═══════════════════════════════════════════════════════════════════════
# HOJA 5: MENSUAL (resumen por mes)
# ═══════════════════════════════════════════════════════════════════════
ws_mes = wb.create_sheet("Resumen Mensual")
ws_mes.sheet_properties.tabColor = "FFC000"

mes_headers = ["Mes", "Ventas Totales", "Costos Totales", "Beneficio Bruto",
               "Margen %", "Cant. Ventas", "Ticket Promedio"]
for i, h in enumerate(mes_headers, 1):
    ws_mes.cell(row=1, column=i, value=h)
style_header(ws_mes, len(mes_headers))

meses = ["Enero 2025", "Febrero 2025", "Marzo 2025", "Abril 2025", "Mayo 2025"]
for i, mes in enumerate(meses):
    r = 2 + i
    ws_mes.cell(row=r, column=1, value=mes).border = THIN_BORDER
    # Placeholder - user fills real data or links to Ventas
    ws_mes.cell(row=r, column=2, value=0).border = THIN_BORDER
    ws_mes.cell(row=r, column=3, value=0).border = THIN_BORDER
    ws_mes.cell(row=r, column=4, value=f'=B{r}-C{r}').border = THIN_BORDER
    ws_mes.cell(row=r, column=5, value=f'=IF(B{r}>0,D{r}/B{r},0)').border = THIN_BORDER
    ws_mes.cell(row=r, column=6, value=0).border = THIN_BORDER
    ws_mes.cell(row=r, column=7, value=f'=IF(F{r}>0,B{r}/F{r},0)').border = THIN_BORDER
    for col in [2, 3, 4, 7]:
        ws_mes.cell(row=r, column=col).number_format = CURRENCY_FMT
    ws_mes.cell(row=r, column=5).number_format = PCT_FMT

# Totales
tr = len(meses) + 2
ws_mes.cell(row=tr, column=1, value="TOTAL ANUAL").font = Font(bold=True)
ws_mes.cell(row=tr, column=1).border = THIN_BORDER
for c in range(2, 8):
    ws_mes.cell(row=tr, column=c, value=f'=SUM({get_column_letter(c)}2:{get_column_letter(c)}{tr-1})')
    ws_mes.cell(row=tr, column=c).font = Font(bold=True)
    ws_mes.cell(row=tr, column=c).border = THIN_BORDER
    if c in (2, 3, 4, 7):
        ws_mes.cell(row=tr, column=c).number_format = CURRENCY_FMT
    elif c == 5:
        ws_mes.cell(row=tr, column=c).value = f'=IF(B{tr}>0,D{tr}/B{tr},0)'
        ws_mes.cell(row=tr, column=c).number_format = PCT_FMT

auto_width(ws_mes, len(mes_headers))

# ── Guardar ──
output = r"C:\Users\maxia\Desktop\omix-code\kiosco_gestion.xlsx"
wb.save(output)
print(f"OK -> {output}")
